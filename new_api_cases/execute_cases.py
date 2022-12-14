# coding:utf-8
import json
import os
import re
import threading
import time
from datetime import datetime
import jsonpath
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from basic_info.mylogging import myLog
from httpop.Httpop import Httpop
from util import get_host
from openpyxl import load_workbook
import requests
from basic_info.get_auth_token import get_headers, get_headers_root,get_auth_token
from util.elasticsearch7 import get_es_data, get_es_data_for_thumbnailMode
from util.encrypt import encrypt_rf
from util.format_res import dict_res, get_time
from basic_info.setting import MySQL_CONFIG, MY_LOGIN_INFO2, baymax_sheet, baymax_master
from util.Open_DB import MYSQL
from basic_info.ready_dataflow_data import get_dataflow_data, get_executions_data, set_upsert_data, query_dataflow_data
from basic_info.setting import tenant_id_83,host
from new_api_cases.deal_parameters import deal_parameters
import unittest
from new_api_cases.get_statementId import statementId_flow_use, statementId_flow_output_use
from new_api_cases.get_statementId import statementId, statementId_no_dataset, get_sql_analyse_statement_id, \
    get_sql_analyse_dataset_info, get_sql_execte_statement_id, steps_sql_parseinit_statemenId, \
    steps_sql_analyzeinit_statementId,get_step_output_init_statementId,get_step_output_ensure_statementId
from new_api_cases.prepare_datas_for_cases import get_job_tasks_id, collector_schema_sync, get_applicationId, \
    get_woven_qaoutput_dataset_path, upload_jar_file_workflow, upload_jar_file_dataflow, upload_jar_file_filter, \
    dss_data, upddss_data, dataset_data, upddataset_data, create_schema_data, updschema_data, create_flow_data, \
    update_flow_data, filesets_data, get_old_id_name, get_collector_data, tag_data, set_user_role, update_role, \
    update_user, enable_user, enable_role

ms = MYSQL(MySQL_CONFIG["HOST"], MySQL_CONFIG["USER"], MySQL_CONFIG["PASSWORD"], MySQL_CONFIG["DB"],MySQL_CONFIG["PORT"])
ab_dir = lambda n: os.path.abspath(os.path.join(os.path.dirname(__file__), n))
case_table = load_workbook(ab_dir("api_cases.xlsx"))
case_table_sheet = case_table.get_sheet_by_name(baymax_master)
all_rows = case_table_sheet.max_row
fileset_dir=os.path.join(os.path.abspath('.'),'attachment\Capture001.png')
log=myLog().getLog().logger
minio_data=[]
httpop=Httpop()
host=host
# ??????????????????????????????????????????????????????????????????????????????
def deal_request_method():
    for i in range(2, all_rows+1):
        request_method = case_table_sheet.cell(row=i, column=4).value
        old_request_url = host+case_table_sheet.cell(row=i, column=5).value
        request_url = deal_parameters(old_request_url)
        # host = get_host.get_host(request_url)
        old_data = case_table_sheet.cell(row=i, column=6).value
        request_data = deal_parameters(old_data)
        #request_data = old_data.encode('utf-8')
        print("request data:", request_data)
        key_word = case_table_sheet.cell(row=i, column=3).value
        api_name = case_table_sheet.cell(row=i, column=1).value
        # ?????????????????????
        if request_method:
            request_method_upper = request_method.upper()
            if api_name == 'tenants':  # ???????????????????????????root?????????????????????
                # ??????????????????????????????????????????
                if request_method_upper == 'POST':
                    # ??????post??????????????????
                    post_request_result_check(row=i, column=8, url=request_url, host=host, headers=get_headers_root(host),
                                              data=request_data, table_sheet_name=case_table_sheet)

                elif request_method_upper == 'GET':
                    # ??????GET??????
                    get_request_result_check(url=request_url, host=host, headers=get_headers_root(host), data=request_data,
                                             table_sheet_name=case_table_sheet, row=i, column=8)

                elif request_method_upper == 'PUT':
                    put_request_result_check(url=request_url, host=host, row=i, data=request_data,
                                             table_sheet_name=case_table_sheet, column=8, headers=get_headers_root(host))

                elif request_method_upper == 'DELETE':
                    delete_request_result_check(request_url, request_data, host=host, table_sheet_name=case_table_sheet, row=i,
                                                column=8, headers=get_headers_root(host))
                else:
                    print('????????????%s?????????????????????' % request_method)
            else:
                # ??????????????????????????????????????????
                if request_method_upper == 'POST':
                    # ??????post??????????????????
                    post_request_result_check(row=i, host=host, column=8, url=request_url, headers=get_headers(host),
                                              data=request_data, table_sheet_name=case_table_sheet)

                elif request_method_upper == 'GET':
                    # ??????GET??????
                    get_request_result_check(url=request_url, host=host, headers=get_headers(host), data=request_data,
                                             table_sheet_name=case_table_sheet, row=i, column=8)

                elif request_method_upper == 'PUT':
                    put_request_result_check(url=request_url, host=host, row=i, data=request_data, table_sheet_name=case_table_sheet, column=8, headers=get_headers(host))

                elif request_method_upper == 'DELETE':
                    delete_request_result_check(url=request_url, host=host, data=request_data,table_sheet_name=case_table_sheet,row=i,column=8, headers=get_headers(host))

                elif request_method_upper == 'TEST':
                    t1=threading.Thread(target=test1, args=(request_url, request_data, host, case_table_sheet, i,8, get_headers(host)))
                    t1.start()

                else:
                    print('????????????%s?????????????????????' % request_method)
        else:
            print('??? %d ?????????????????????' % i)
    #  ???????????????????????????
    case_table.save(ab_dir("api_cases.xlsx"))

def test1(url, data, host, table_sheet_name,row, column,headers):
    global response
    count = 0
    new_data={"fieldList":[{"logicalOperator":"AND","fieldName":"name","comparatorOperator":"LIKE","fieldValue":'%'+data+'%'},{"logicalOperator":"AND","fieldName":"flowType","comparatorOperator":"EQUAL","fieldValue":"dataflow"},{"logicalOperator":"AND","fieldName":"flowId","comparatorOperator":"EQUAL","fieldValue":"9b2b25fe-29e1-4874-abf6-d7741c091848"}],"sortObject":{"field":"lastModifiedTime","orderDirection":"DESC"},"offset":0,"limit":8}
    new_data=json.dumps(new_data)
    print(new_data)
    while count <=10:
        print(url,new_data,headers)
        response = httpop.api_post(url=url, headers=headers, data=new_data)
        print("response",response)
        print("response.text",response.text)
        response_new=json.loads(response.text)
        print("response_new",response_new)
        if response_new["content"]==[]:
            print("sleep",datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            time.sleep(60)
            break
        count+=1
    print("???????????????",response.text)
    clean_vaule(table_sheet_name, row, column)
    write_result(sheet=table_sheet_name, row=row, column=column, value= response.status_code)
    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    case_table.save(ab_dir("api_cases.xlsx"))


# POST??????
def post_request_result_check(row, column, url, host, headers, data, table_sheet_name):
    global case_detail
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("???????????????%s" % case_detail)
        if '(Id?????????)' in case_detail:
            # ?????????statementId,???????????????URL??????????????????
            print('???????????????', case_detail)
            statement = statementId_no_dataset(host, dict_res(data))
            new_url = url.format(statement)
            data = data.encode('utf-8')
            response=httpop.api_post(url=new_url, headers=headers, data=data)
            print(response.text, response.status_code)
            # ????????????status_code???response.text???????????????10?????????14???
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '???????????????' in case_detail:
            # ?????????statementId,???????????????URL??????????????????
            print('???????????????', case_detail)
            print("new_url:", url)
            new_data = dss_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data:", new_data)
            # response = httpop.api_post(url=url, headers=headers, data=new_data)
            response=httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.text, response.status_code)
            # print(response.url)
            # ????????????status_code???response.text???????????????10?????????14???
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '??????name???????????????' in case_detail:
            print('???????????????', case_detail)
            new_data = dss_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data:", new_data)
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.text, response.status_code)
            # print(response.url)
            # ????????????status_code???response.text???????????????10?????????14???
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '??????schema' in case_detail:
            # ?????????statementId,???????????????URL??????????????????
            print('???????????????', case_detail)
            new_data = create_schema_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data:", new_data)
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.text, response.status_code)
            # print(response.url)
            # ????????????status_code???response.text???????????????10?????????14???
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '??????flow' in case_detail:
            # ?????????statementId,???????????????URL??????????????????
            print('???????????????', case_detail)
            new_data = create_flow_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data:", new_data)
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.text, response.status_code)
            # ????????????status_code???response.text???????????????10?????????14???
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '?????????????????????':
            # ?????????statementId,???????????????URL??????????????????
            print('???????????????', case_detail)
            new_data = dss_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data:", new_data)
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.text, response.status_code)
            # ????????????status_code???response.text???????????????10?????????14???
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????JDBC???????????????':
            print('???????????????', case_detail)
            dss_id, new_data = upddss_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data:", new_data)
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(table_sheet_name, row, column, response.status_code)
            write_result(table_sheet_name, row, column+4, response.text)
        elif 'datasetId?????????'in case_detail:
            # ?????????statementId,???????????????URL??????????????????
            print('???????????????', case_detail)
            new_data = dataset_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data:", new_data)
            response = httpop.api_get(url=url, headers=headers)
            print(response.text, response.status_code)
            # ????????????status_code???response.text???????????????10?????????14???
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)

        elif case_detail == '??????statementId???Dataset??????(no datasetId)':
            # ?????????statementId,???????????????URL??????????????????
            print('???????????????', case_detail)
            statement = statementId_no_dataset(host, dict_res(data))
            print("statementid:", statement)
            new_url = url.format(statement)
            data = data.encode('utf-8')
            response = httpop.api_get(url=new_url, headers=headers)
            print(response.text, response.status_code)
            # print(response.url)
            # ????????????status_code???response.text???????????????10?????????14???
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '???Dataset????????????,??????statement  id(datasetId??????)':
            # ?????????statementId,???????????????URL??????????????????
            print('???????????????', case_detail)
            new_data = dataset_data(data)
            #new_url = url.format(dataset_id)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data1:", new_data)
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            # ????????????status_code???response.text???????????????10?????????14???
            print(response.text, response.status_code)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????statement id,????????????Dataset???????????????(datasetId??????)':
            # ?????????statementId,???????????????URL??????????????????
            print('???????????????', case_detail)
            statement_id, new_data = statementId(host, data)
            new_url = url.format(statement_id)
            print(new_url)
            #new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data2:", new_data)
            res = httpop.api_post(url=new_url, headers=headers,data=new_data)
            count_num = 0
            while ("waiting") in res.text or ("running") in res.text:
                print('???????????????', res.text)
                res = requests.post(url=new_url, headers=headers,data=new_data)
                count_num += 1
                if count_num == 100:
                    return
                print('???????????????', res.text)
            # ????????????str??????
            print(res.text)
            if '"statement":"available"' in res.text:
             #????????????status_code???response.text???????????????10?????????14???
                print(res.text, res.status_code)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=res.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=res.text)
            else:
                print(res.text, res.status_code)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=res.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=res.text)
        elif case_detail == '??????SQL??????????????????':
            print('???????????????', case_detail)
            # ??????????????????????????????statement id ??? ?????????????????????
            execte_use_params = get_sql_analyse_dataset_info(host, data)  # ?????????????????????
            execte_use_params = json.dumps(execte_use_params, separators=(',', ':'))
            execte_statement_id = get_sql_execte_statement_id(host, data)  # statement id
            new_url = url.format(execte_statement_id)
            response = httpop.api_post(url=new_url, headers=headers, data=execte_use_params)
            count_num = 0
            while ("waiting") in response.text or ("running") in response.text:
                print('???????????????', response.text)
                response = httpop.api_post(url=new_url, headers=headers, data=execte_use_params)
                count_num += 1
                if count_num == 100:
                    return
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            # print(response.status_code)
            # print(response.text)
        elif '??????dataset'in case_detail:
            print('???????????????', case_detail)
            new_data = dataset_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data:", new_data)
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.text, response.status_code)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '????????????':
            log.info("request   url???%s" %url)
            new_data = update_user(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data???%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '????????????':
            log.info("request   url???%s" %url)
            new_data = update_role(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data: ",new_data)
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data???%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????????????????':
            new_datas = set_user_role(data)
            new_url = url.format(new_datas[0])
            log.info("new_url???%s"% new_url)
            new_data = json.dumps(new_datas[1], separators=(',', ':'))
            response = requests.post(url=new_url, data=new_data, headers=headers)
            log.info("response data???%s %s"%(response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '?????????????????????':
            new_datas = set_user_role(data)
            new_url = url.format(new_datas[0])
            log.info("new_url???%s"% new_url)
            new_data = json.dumps(new_datas[2], separators=(',', ':'))
            response = requests.post(url=new_url, data=new_data, headers=headers)
            log.info("response data???%s %s"%(response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????es-upsert-dataflow':
            print('???????????????', case_detail)
            #insert???update training
            set_upsert_data()
            new_data = get_dataflow_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = httpop.api_post(url=url, headers=headers, data=data)
            print(response.text, response.status_code)
            count_num = 0
            while ("waiting") in response.text or ("running") in response.text:
                print('???????????????',response.text)
                response = httpop.api_post(url=url, headers=headers, data=data)
                count_num += 1
                if count_num == 100:
                    return
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????colSplit-pivot-unpivot-Explode-dataflow':
            print('???????????????', case_detail)
            new_data = get_dataflow_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.text, response.status_code)
            count_num = 0
            while ("waiting") in response.text or ("RUNNING") in response.text:
                print('???????????????', response.text)
                response = httpop.api_post(url=url, headers=headers, data=data)
                time.sleep(5)
                count_num += 1
                if count_num == 100:
                    return
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????executions-colSplit-pivot-unpivot-Explode-dataflow':
            print('???????????????', case_detail)
            new_data = query_dataflow_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.text, response.status_code)
            count_num = 0
            time.sleep(5)
            while ("waiting") in response.text or ("READY") in response.text or ("RUNNING") in response.text:
                print('???????????????', response.text)
                response = httpop.api_post(url=url, headers=headers, data=new_data)
                time.sleep(5)
                count_num += 1
                if count_num == 150:
                    return
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????executionOutputs-executions-colSplit-pivot-unpivot-ExplodeStep-dataflow':
            print('???????????????', case_detail)
            new_data = get_executions_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.text, response.status_code)
            count_num = 0
            while ("waiting") in response.text or ("READY") in response.text or ("RUNNING") in response.text:
                print('???????????????', response.text)
                response = httpop.api_post(url=url, headers=headers, data=data)
                count_num += 1
                if count_num == 10:
                    return
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????executions-es-upsert-dataflow':
            print('???????????????', case_detail)
            new_data = query_dataflow_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("url:", url)
            response = httpop.api_post(url=url, headers=headers, data=data)
            count_num = 0
            time.sleep(5)
            while ("waiting") in response.text or ("READY") in response.text or ("RUNNING") in response.text:
                #print('???????????????', response.text)
                response = httpop.api_post(url=url, headers=headers, data=data)
                time.sleep(5)
                count_num += 1
                #if ('"type":"SUCCEEDED"') in response.text or ("FAILED")in response.text or ("KILLED") in response.text:
                #return
                if count_num == 80:
                    return
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????executionOutputs-es-upsert-dataflow':
            print('???????????????', case_detail)
            new_data = get_executions_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            count_num = 0
            while ("waiting") in response.text or ("READY") in response.text or ("RUNNING") in response.text:
                print('???????????????', response.text)
                response = httpop.api_post(url=url, headers=headers, data=data)
                count_num += 1
                if count_num == 10:
                    return
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '????????????execution':
            print('???????????????', case_detail)
            # ?????????????????????flow????????????execution???????????????execution id????????????list????????????????????????
            # query_execution_url = '%s/api/executions/query' % host
            # all_exectuions = httpop.api_post(url=query_execution_url, headers=headers, data=data)
            # executions_dict = dict_res(all_exectuions.text)
            # try:
            #     executions_content = executions_dict['content']
            #     all_ids = [] # ???list?????????????????????execution id
            #     for item in executions_content:
            #         executions_content_id = item['id']
            #         all_ids.append(executions_content_id)
            # except Exception as e:
            #     print(e)
            # else:  # ????????????id??????????????????list??????????????????removeLIst???????????????
            #     removelist_data = []
            #     removelist_data.append(all_ids[-1])
            #     # ??????????????????
            #     removeList_result = httpop.api_post(url=url, headers=headers, json=removelist_data)
            #     print(removeList_result.text, removeList_result.status_code)
            data=json.dumps(data, separators=(',', ':'))
            response = httpop.api_post(url=url, headers=headers, data=data)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)

        elif case_detail == '????????????????????????????????????':
            print('???????????????', case_detail)
            task_id = get_job_tasks_id(data)
            response = httpop.api_post(url=url, headers=headers, json=task_id)
            print(response.text, response.status_code)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '??????schema'in case_detail:
            print('???????????????', case_detail)
            print("data:", data)
            new_data = json.dumps(data, separators=(',', ':'))
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.text, response.status_code)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '??????Sql??????????????????,?????????ParseSql??????'in case_detail:
            print('???????????????', case_detail)
            print("data:", data)
            #new_data = json.dumps(data, separators=(',', ':'))
            response = httpop.api_post(url=url, headers=headers, data=data)
            print(response.text, response.status_code)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '?????????Sql'in case_detail:
            print('???????????????', case_detail)
            print("data:", data)
            #new_data = json.dumps(data, separators=(',', ':'))
            response = httpop.api_post(url=url, headers=headers, data=data)
            print(response.text, response.status_code)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????????????????????????????':
            print('???????????????', case_detail)
            response = httpop.api_post(url=url, headers=headers, json=dict_res(data))
            print(response.text, response.status_code)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????????????????????????????dataset':
            print('???????????????', case_detail)
            response = httpop.api_post(url=url, headers=headers, data=data)
            print(response.text, response.status_code)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '????????????????????????????????????':  # ???????????????????????????????????????????????????
            data = {"fieldList":[{"fieldName":"createTime","fieldValue":get_time(),"comparatorOperator":"GREATER_THAN","logicalOperator":"AND"},{"fieldName":"createTime","fieldValue":1555516800000,"comparatorOperator":"LESS_THAN"}],"sortObject":{"field":"lastModifiedTime","orderDirection":"DESC"},"offset":0,"limit":8,"groupBy":"testTime"}
            response = httpop.api_post(url=url,headers=headers, json=data)
            print(response.status_code,response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail in ('????????????????????????-??????jar???', '???????????????-??????jar???', '????????????????????????-??????jar???'):
            print('???????????????', case_detail)
            files = {"file": open(fileset_dir, 'rb')}
            headers.pop('Content-Type')
            response = httpop.api_post(url=url, files=files, headers=headers)
            print(response.text, response.status_code)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '???????????????5????????????':
            print('???????????????', case_detail)
            new_data = json.dumps(data, separators=(',', ':'))
            print(new_data)
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.text, response.status_code)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????statement id,????????????Dataset???????????????(datasetId??????)':
            print('???????????????', case_detail)
            statementId1, new_data = statementId(host, data)
            print("fh",statementId1, new_data)
            new_url = url.format(statementId1)
            response = httpop.api_post(url=new_url, headers=headers, data=new_data)
            print(response.text, response.status_code)
            count_num = 0
            while "running" in response.text or "waiting" in response.text:
                time.sleep(5)
                response = httpop.api_post(url=new_url, headers=headers, data=new_data)
                count_num += 1
                if count_num == 100:
                    return
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        # elif case_detail == '????????????????????????':
        #     fileName = upload_jar_file_workflow()
        #     new_url = url.format(fileName)
        #     response = httpop.api_post(url=new_url, headers=headers, data=data)
        #     print(response.text, response.status_code)
        #     clean_vaule(table_sheet_name, row, column)
        #     write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
        #     write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        # elif case_detail == '????????????????????????':
        #     fileName = upload_jar_file_dataflow()
        #     new_url = url.format(fileName)
        #     response = httpop.api_post(url=new_url, headers=headers, data=data)
        #     print(response.text, response.status_code)
        #     clean_vaule(table_sheet_name, row, column)
        #     write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
        #     write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        # elif case_detail == '???????????????':
        #     fileName = upload_jar_file_filter()
        #     new_url = url.format(fileName)
        #     response = httpop.api_post(url=new_url, headers=headers, data=data)
        #     print(response.text, response.status_code)
        #     clean_vaule(table_sheet_name, row, column)
        #     write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
        #     write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '????????????????????????':
            dir2 = ab_dir('sex.xls')
            files = {"file": open(dir2, 'rb')}
            headers = get_headers(host)
            headers.pop('Content-Type')
            response = httpop.api_post(url, files=files, headers=headers)
            print(response.text, response.status_code)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)

        elif case_detail == "????????????":
            headers["Content-Type"] = "application/x-www-form-urlencoded"
            headers.pop('X-AUTH-TOKEN')
            response = httpop.api_post(url, headers=headers, data=dict_res(data))
            print(response.text, response.status_code)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '??????????????????????????????':
            try:
                print(data)
                new_data = {"fieldList": [
                    {"fieldName": "executionId", "fieldValue": data, "comparatorOperator": "EQUAL",
                     "logicalOperator": "AND"}], "sortObject": {"field": "lastModifiedTime", "orderDirection": "DESC"},
                    "offset": 0, "limit": 8}
                print(new_data)
                response = httpop.api_post(url=url, headers=headers, json=new_data)
                print(response.text, response.status_code)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            except Exception:
                print('????????????flow???tc_df_top_???????????? ???????????????execution id')
        elif case_detail == '???????????????????????????':
            # ????????????id???user
            new_data = [{"id": data, "enabled": 0}]
            del_url = '%s/api/user/resetStatus' % host
            res = httpop.api_post(url=del_url, headers=headers, json=new_data)
            # ????????????id???user
            del_user_id = []
            del_user_id.append(data)
            response = httpop.api_post(url=url, headers=headers, json=del_user_id)
            print('????????????', response.status_code, response.content)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '??????' in case_detail :
            new_headers = {'Content-Type': 'application/x-www-form-urlencoded'}
            if case_detail == '??????':
                data = {'name': encrypt_rf('admin'), 'password': encrypt_rf('123456'), 'version': 'Europa-3.0.0.19 - 20180428', 'tenant': encrypt_rf('default')}
                response = httpop.api_post(url=url, headers=new_headers, data=data)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '???????????????????????????':
                data = {'name': encrypt_rf('admin'), 'password': encrypt_rf('123456555'), 'version': 'Europa-3.0.0.19 - 20180428', 'tenant': encrypt_rf('default')}
                response = httpop.api_post(url=url, headers=new_headers, data=data)
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '????????????????????????':
                data = {'name': encrypt_rf('admin12399999999999999'), 'password': encrypt_rf('123456'), 'version': 'Europa-3.0.0.19 - 20180428', 'tenant': encrypt_rf('default')}
                response = httpop.api_post(url=url, headers=new_headers, data=data)
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '???????????????????????????':
                data = {'name': encrypt_rf('user_without_pression'), 'password': encrypt_rf('123456'),
                        'version': 'Europa-3.0.0.19 - 20180428', 'tenant': encrypt_rf('default')}
                response = httpop.api_post(url=url, headers=new_headers, data=data)
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '???????????????????????????????????????':
                data = {'name': encrypt_rf('user_pwd_expired'), 'password': encrypt_rf('123456'),
                        'version': 'Europa-3.0.0.19 - 20180428', 'tenant': encrypt_rf('default')}
                response = httpop.api_post(url=url, headers=new_headers, data=data)
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '????????????????????????????????????????????????':
                data = {'name': encrypt_rf('user_time_expired'), 'password': encrypt_rf('123456'),
                        'version': 'Europa-3.0.0.19 - 20180428', 'tenant': encrypt_rf('default')}
                response = httpop.api_post(url=url, headers=new_headers, data=data)
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '????????????????????????????????????????????????????????????':
                data = {'name': encrypt_rf('user_expired'), 'password': encrypt_rf('123456'),
                        'version': 'Europa-3.0.0.19 - 20180428', 'tenant': encrypt_rf('default')}
                response = httpop.api_post(url=url, headers=new_headers, data=data)
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '??????????????????????????????' in case_detail:
            new_data=filesets_data(data)
            # new_data = json.dumps(new_data, separators=(',', ':'))
            response=httpop.api_post(url=url, headers=headers, json=new_data)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '??????flow' in case_detail:
            new_data=get_old_id_name(data)
            # new_data = json.dumps(new_data, separators=(',', ':'))
            response=httpop.api_post(url=url, headers=headers, json=new_data)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '????????????????????????' in case_detail:
            new_data=json.dumps(data)
            response=httpop.api_post(url=url, headers=headers, data=new_data)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '???????????????????????????????????????' in case_detail:
            new_url=url.format(data)
            data=get_collector_data(data)
            response = httpop.api_post(url=new_url, headers=headers, data=data)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(table_sheet_name, row, column, response.status_code)
            write_result(table_sheet_name, row, column+4, response.text)
        elif ('??????ID??????????????????????????????' or '????????????job id???task') in case_detail:
            new_url=url.format(data)
            data={"fieldList":[],"sortObject":{"field":"lastModifiedTime","orderDirection":"DESC"},"offset":0,"limit":8}
            response = httpop.api_post(url=new_url, headers=headers, json=data)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(table_sheet_name, row, column, response.status_code)
            write_result(table_sheet_name, row, column+4, response.text)
        elif ('????????????????????????') in case_detail:
            para=data.split('&')
            es_id=get_es_data(para[0],para[1],para[2],eval(para[3]))
            content=para[4]
            new_data={"content":content,"offset":0,"limit":8,"ids":es_id}
            response = httpop.api_post(url=url, headers=headers, json=new_data)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(table_sheet_name, row, column, response.status_code)
            write_result(table_sheet_name, row, column+4, response.text)
        elif "??????ES????????????" in case_detail:
            para=data.split("&")
            es_id=get_es_data(para[0],para[1],para[2],eval(para[3]))
            new_data=json.dumps(es_id)
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif ("????????????" or "????????????") in case_detail:
            para=data.split("&")
            es_id=get_es_data(para[0],para[1],para[2],eval(para[3]))
            data={"ids":es_id,"tags":eval(para[4])}
            # new_data=json.dumps(es_id)
            response = httpop.api_post(url=url, headers=headers, json=data)
            print(response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif ("????????????") in case_detail:
            para=data.split("&")
            es_id=get_es_data(para[0],para[1],para[2],eval(para[3]))
            data={"ids":es_id}
            # new_data=json.dumps(es_id)
            response = httpop.api_post(url=url, headers=headers, json=data)
            print(response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif case_detail=="????????????":
            para=data.split("&")
            es_id=get_es_data(para[0],para[1],para[2],eval(para[3]))
            new_data=json.dumps(es_id)
            response = httpop.api_post(url=url, headers=headers, data=new_data)
            print(response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif "???????????????" in case_detail:
            para=data.split("&")
            es_id=get_es_data(para[0],para[1],para[2],eval(para[3]))
            new_url=url.format(es_id[0])
            # new_data=json.dumps(es_id)
            response = httpop.api_post(url=new_url, headers=headers, json=data)
            print(response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif "????????????" in case_detail:
            para=data.split("&")
            es_id=get_es_data(para[0],para[1],para[2],eval(para[3]))
            new_url=url.format(es_id[0])
            fs = {"file": open(fileset_dir, 'rb')}
            headers.pop('Content-Type')
            headers["Accept"]='*/*'
            response = httpop.api_post(url=new_url, headers=headers, files=fs)
            print(response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif ("??????????????????") in case_detail:
            para=data.split("&")
            es_ids=get_es_data_for_thumbnailMode(para[0],para[1],para[2])
            es_ids=json.dumps(es_ids)
            response = httpop.api_post(url=url, headers=headers, data=es_ids)
            print(response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif ("????????????????????????") in case_detail:
            new_url=url.format(data)
            minio_data.append(data)
            if "MINIO" in case_detail:
                data={"password":"inforefiner","port":"9000","host":"192.168.1.81","region":"","username":"minio"}
            elif "OZONE" in case_detail:
                data= {}
            response = httpop.api_post(url=new_url, headers=headers, json=data)
            print(response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif "????????????_MINIO" in case_detail:
            # new_url=url.format('/'+minio_data[0])
            fs = {"file": open(fileset_dir, 'rb')}
            headers.pop('Content-Type')
            response = httpop.api_post(url=url, headers=headers, files=fs)
            print(response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif "????????????_OZONE" in case_detail:
            fs = {"file": open(fileset_dir, 'rb')}
            headers.pop('Content-Type')
            response = httpop.api_post(url=url, headers=headers, files=fs)
            print(response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))



        else:
            print('???????????????', case_detail)
            if data:
                data = str(data)
                # ??????????????????????????????{"id":"7135cf6e-2b12-4282-90c4-bed9e2097d57","name":"gbj_for_jdbcDatasource_create_0301_1_0688","creator":"admin"}
                if '&' in data:
                    # ????????????
                    parameters = data.split('&')
                    # ??????URL
                    new_url = url.format(parameters[0])
                    # ??????????????????
                    parameters_data = parameters[-1]
                    if parameters_data.startswith('{'):
                        response = httpop.api_post(url=new_url, headers=headers, json=dict_res(parameters_data))
                        print("response data:", response.status_code, response.text)
                        clean_vaule(table_sheet_name, row, column)
                        write_result(table_sheet_name, row, column, response.status_code)
                        write_result(table_sheet_name, row, column+4, response.text)
                    else:
                        print('????????????%d???parameters?????????update?????????????????????id&{data}' % row)
                elif data.startswith('select id'):
                    result = ms.ExecuQuery(data)
                    if result:
                        new_data = result[0]["id"]
                        new_url = url.format(new_data)
                        response = httpop.api_post(url=url, headers=headers, data=new_data)
                        print("response data:", response.status_code, response.text)
                        clean_vaule(table_sheet_name, row, column)
                        write_result(table_sheet_name, row, column, response.status_code)
                        write_result(table_sheet_name, row, column + 4, response.text)
                    else:
                        print('?????????result:???', result)
                elif data.startswith('{') and data.endswith('}'):
                    data_dict = dict_res(data)
                    response = httpop.api_post(url=url, headers=headers, json=data_dict)
                    print("response data:", response.status_code, response.text)
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                # ????????????????????? ???["9d3639f0-02bc-44cd-ac71-9a6d0f572632"]
                elif data.startswith('[') and data.endswith(']'):
                    if "'" in data:
                        data = data.replace("'", '"')
                        print(data)
                        response = httpop.api_post(url=url, headers=headers, data=data)
                        print("response data:", response.status_code, response.text)
                        clean_vaule(table_sheet_name, row, column)
                        write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                        write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                    else:
                        print(data)
                        response = httpop.api_post(url=url, headers=headers, data=data)
                        print("response data:", response.status_code, response.text)
                        clean_vaule(table_sheet_name, row, column)
                        write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                        write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                else:  # ??????????????????????????????id?????????,???????????????????????????list?????????
                    new_data = []
                    new_data.append(data)
                    new_data = str(new_data)
                    if "'" in new_data:
                        print(data)
                        new_data = new_data.replace("'", '"')
                        response = httpop.api_post(url=url, headers=headers, data=new_data)
                        print("response data:", response.status_code, response.text)
                        clean_vaule(table_sheet_name, row, column)
                        write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                        write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                    else:
                        print(data)
                        response = httpop.api_post(url=url, headers=headers, data=new_data)
                        print("response data:", response.status_code, response.text)
                        clean_vaule(table_sheet_name, row, column)
                        write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                        write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            else:
                print(data)
                response = httpop.api_post(url=url, headers=headers, data=data)
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    except Exception as e:
        log.error("????????????{}?????????????????????{}".format(case_detail,e))
        clean_vaule(table_sheet_name, row, column)
        write_result(sheet=table_sheet_name, row=row, column=column, value='-1')
        write_result(sheet=table_sheet_name, row=row, column=column + 4, value='{"id":"-1"}')


    # else:
    #     print('????????????%d??????data??????' % row)


# GET??????
def get_request_result_check(url, headers, host, data, table_sheet_name, row, column):
    case_detail = case_table_sheet.cell(row=row, column=2).value
    log.info("???????????????%s" % case_detail)
    # GET???????????????parameter???????????????,?????????????????????URL??????
    try:
        if data:
            if '(Id??????)' in case_detail:
                # print(data)
                print('???????????????', case_detail)
                # data = deal_parameters(data)
                statement_id = statementId(host, data)
                parameter_list = []
                parameter_list.append(data)
                parameter_list.append(statement_id)
                url_new = url.format(parameter_list[0], parameter_list[1])
                response = httpop.api_get(url=url_new, headers=headers)
                print(response.text, response.status_code)
                count_num = 0
                while response.text in ('{"statement":"waiting"}', '{"statement":"running"}'):
                    response = httpop.api_get(url=url_new, headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == ('??????statement id,??????Sql Analyze??????(??????????????????)'):
                print('???????????????', case_detail)
                sql_analyse_statement_id = get_sql_analyse_statement_id(host, data)
                new_url = url.format(sql_analyse_statement_id)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == ('????????????statementId?????????????????????'):  # ??????SQL analyse??????
                print('???????????????', case_detail)
                cancel_statement_id = get_sql_analyse_statement_id(host, data)
                new_url = url.format(cancel_statement_id)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                clean_vaule(table_sheet_name,
                            row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)

            elif case_detail == ('????????????sql parse???????????????statementId,??????dataset name'):
                print('???????????????', case_detail)
                data = data.encode('utf-8')
                datasetName_statementId = steps_sql_parseinit_statemenId(host, data)
                new_url = url.format(datasetName_statementId)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                count_num = 0
                while response.text in ('{"statement":"waiting"}', '{"statement":"running"}'):
                    response = httpop.api_get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == ('??????Sql Analyze?????????statementId,??????SqlAnalyze??????'):
                print('???????????????', case_detail)
                steps_sql_analyse_statementId = steps_sql_analyzeinit_statementId(host, data)
                new_url = url.format(steps_sql_analyse_statementId)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                count_num = 0
                while "waiting" in response.text or "running"in response.text:
                    response = httpop.api_get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == ('??????sqlsource step?????????statementId????????????'):
                print('???????????????', case_detail)
                data = data.encode('utf-8')
                cancel_sql_parseinit_statementId = steps_sql_parseinit_statemenId(host, data)
                new_url = url.format(cancel_sql_parseinit_statementId)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail in ('??????????????????????????????????????????','????????????????????????????????????','??????tasks id ????????????log'):
                print('???????????????', case_detail)
                time.sleep(10)
                task_id = collector_schema_sync(data)
                time.sleep(5)
                new_url = url.format(task_id)
                # time.sleep(2)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '??????flow':
                print('???????????????', case_detail)
                token = get_auth_token(host)
                new_url = url.format(token)
                response = httpop.api_get(url=new_url,headers=headers)
                print(response.text, response.status_code)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '????????????id?????????':
                print('???????????????', case_detail)
                new_url = url.format(data)
                print(new_url)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '??????statementID??????step???????????????':
                print('???????????????', case_detail)
                init_statementId = get_step_output_init_statementId(host, data)
                # print(init_statementId)
                new_url = url.format(init_statementId)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                count_num = 1
                while "running" in response.text or "waiting" in response.text:
                    time.sleep(5)
                    response = httpop.api_get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '??????statementID??????step':
                print('???????????????', case_detail)
                ensure_statementId = get_step_output_ensure_statementId(host, data)
                new_url = url.format(ensure_statementId)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                count_num = 0
                while "running" in response.text or "waiting" in response.text:
                    time.sleep(5)
                    response = httpop.api_get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)

            elif case_detail == '??????statement id,????????????colsplit-Dataset???????????????':
                print('???????????????', case_detail)
                res_statementId = statementId_flow_output_use(host, data)
                new_url = url.format(data, res_statementId)
                print("new_url: ", new_url)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                count_num = 0
                while "running" in response.text or "waiting" in response.text:
                    time.sleep(5)
                    response = httpop.api_get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '??????statement id,????????????unpovit3-Dataset???????????????':
                print('???????????????', case_detail)
                res_statementId = statementId_flow_output_use(host, data)
                new_url = url.format(data, res_statementId)
                print("new_url: ", new_url)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                count_num = 0
                while "running" in response.text or "waiting" in response.text:
                    time.sleep(5)
                    response = httpop.api_get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                # print(response.url, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '??????statement id,????????????pivot-Dataset???????????????':
                print('???????????????', case_detail)
                res_statementId = statementId_flow_output_use(host, data)
                new_url = url.format(data, res_statementId)
                print("new_url: ", new_url)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                count_num = 0
                while "running" in response.text or "waiting" in response.text:
                    time.sleep(5)
                    response = httpop.api_get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '??????statement id,????????????explode-Dataset???????????????':
                print('???????????????', case_detail)
                res_statementId = statementId_flow_output_use(host, data)
                new_url = url.format(data, res_statementId)
                print("new_url: ", new_url)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                count_num = 0
                while "running" in response.text or "waiting" in response.text:
                    time.sleep(5)
                    response = httpop.api_get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '??????statement id,????????????es-upsert-dataset???????????????':
                print('???????????????', case_detail)
                res_statementId = statementId_flow_output_use(host, data)
                new_url = url.format(res_statementId)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                count_num = 0
                while "running" in response.text or "waiting" in response.text:
                    time.sleep(5)
                    response = httpop.api_get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == "??????students_info_zmode_????????????????????????????????????????????????":
                print(host, data)
                statement_id = statementId_flow_use(host, data,tenant_id_83)
                url = url.format(data,statement_id)
                response = httpop.api_get(url=url,headers=headers)
                print(response.text, response.status_code)
                count_num = 0
                while ("waiting") in response.text or ("running") in response.text:
                    print('???????????????',response.status_code, response.text)
                    response = response = httpop.api_get(url=url,headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif "????????????" in case_detail:
                new_url=url.format(data)
                response = httpop.api_get(url=new_url,headers=headers)
                status=json.loads(response.text)["statusType"]
                while status == "WAITTING" or status =="RUNNING" or status =="READY":
                      log.info("------??????while??????------\n")
                      response = httpop.api_get(url=new_url,headers=headers)
                      status=json.loads(response.text)["statusType"]
                      log.info("------???????????????????????????: %s------\n" % status)
                      time.sleep(10)
                if status == "SUCCEEDED":
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                else:
                    print("flow??????????????????")
                    return
            elif 'datasetId?????????'in case_detail:
                # ?????????statementId,???????????????URL??????????????????
                print('???????????????', case_detail)
                # new_data = dataset_data(data)
                # new_data = json.dumps(new_data, separators=(',', ':'))
                # print("new_data:", data)
                response = httpop.api_get(url=url, headers=headers)
                print(response.text, response.status_code)
                # ????????????status_code???response.text???????????????10?????????14???
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif '??????statementId???Dataset??????'in case_detail:
                # ?????????statementId,???????????????URL??????????????????
                print('???????????????', case_detail)
                statement_id, new_data = statementId_no_dataset(host, data)
                #new_data = json.dumps(new_data, separators=(',', ':'))
                new_url = url.format(statement_id)
                print("new_url-new_data:", new_url, new_data)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                # ????????????status_code???response.text???????????????10?????????14???
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif '????????????'in case_detail:
                # ?????????statementId,???????????????URL??????????????????
                print('???????????????', case_detail)
                new_url = url.format(data)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.text, response.status_code)
                # ????????????status_code???response.text???????????????10?????????14???
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            else:
                # print('???????????????', case_detail)
                log.info("????????????{}".format(case_detail))
                if '&' in str(data):  # ????????????????????????&??????
                    parameters = data.split('&')
                    # ????????????select????????????????????????????????????
                    for i in range(len(parameters)):
                        if parameters[i].startswith('select id from'):
                            # select_result = ms.ExecuQuery(parameters[i])
                            try:
                                select_result = ms.ExecuQuery(parameters[i])
                                parameters[i] = select_result[0]["id"]
                            except:
                                print('???%s???????????????????????????' % row)

                        elif parameters[i].startswith('select name from'):
                            try:
                                select_result = ms.ExecuQuery(parameters[i])
                                parameters[i] = select_result[0]["name"]
                            except:
                                print('???%s???????????????????????????' % row)
                        elif parameters[i].startswith('select execution_id from'):
                            try:
                                select_result = ms.ExecuQuery(parameters[i])
                                parameters[i] = select_result[0]["execution_id"]
                            except:
                                print('???%s???????????????????????????' % row)
                    # ??????URL???????????????????????????????????????data??????????????????????????????
                    if len(parameters) == 1:
                        try:
                            url_new = url.format(parameters[0])
                            response = httpop.api_get(url=url_new, headers=headers)
                            print(response.content, response.status_code, response.text)
                        except Exception:
                            return
                        print(response.url, response.status_code,response.text)
                        clean_vaule(table_sheet_name, row, column)
                        write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                        write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                    elif len(parameters) == 2:
                        url_new = url.format(parameters[0], parameters[1])
                        response = httpop.api_get(url=url_new, headers=headers)
                        print("response data:", response.status_code, response.text)
                        clean_vaule(table_sheet_name, row, column)
                        write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                        write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                    elif len(parameters) == 3:
                        url_new = url.format(parameters[0], parameters[1], parameters[2])
                        response = httpop.api_get(url=url_new, headers=headers)
                        print("response data:", response.status_code, response.text)
                        clean_vaule(table_sheet_name, row, column)
                        write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                        write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                    else:
                        print('????????????%d???parameters' % row)
                else:  # ??????????????????&?????????????????????
                    parameters = data
                    url_new = url.format(data)
                    print(url_new)
                    response = httpop.api_get(url=url_new, headers=headers)
                    print("response data:", response.status_code, response.text)
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        # GET ??????????????????URL????????????????????????
        else:
            if case_detail in('??????applicationId??????yarnAppliction??????????????????','??????applicationId??????yarnAppliction???????????????command line log'):
                print('???????????????', case_detail)
                application_id = get_applicationId()
                new_url = url.format(application_id)
                response = httpop.api_get(url=new_url, headers=headers)
                print(response.status_code, response.text, type(response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '????????????????????????path??????dataset-??????datasetId':
                print('???????????????', case_detail)
                dataset_path = get_woven_qaoutput_dataset_path()[0]
                new_url = url.format(dataset_path)
                print(new_url)
                response = httpop.api_get(url=new_url, headers=headers)
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '????????????':
                res = httpop.api_post(url=MY_LOGIN_INFO2["URL"], headers=MY_LOGIN_INFO2["HEADERS"],
                                      data=MY_LOGIN_INFO2["DATA"])
                login_info = dict_res(res.text)
                token = login_info["content"]["access_token"]
                new_url = url.format(token)
                print(new_url)
                response = httpop.api_get(url=new_url,headers=headers)
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            else:
                print('???????????????', case_detail)
                response = httpop.api_get(url=url, headers=headers)
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    except Exception as e:
        log.error("{}?????????????????????{}".format(case_detail, e))
        clean_vaule(table_sheet_name, row, column)
        write_result(sheet=table_sheet_name, row=row, column=column, value='-1')
        write_result(sheet=table_sheet_name, row=row, column=column + 4, value='{"id":"-1"}')


# PUT??????
def put_request_result_check(url, host, row, data, table_sheet_name, column, headers):
    case_detail = case_table_sheet.cell(row=row, column=2).value
    log.info("???????????????%s" % case_detail)
    #if data and isinstance(data, str):
    try:
        if case_detail == '??????????????????':
            print('???????????????', case_detail)
            print(type(data))
            response = requests.put(url=url, headers=headers, data=data)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(table_sheet_name, row, column, response.status_code)
            write_result(table_sheet_name, row, column+4, response.text)
        elif case_detail == '??????schema':
            print('???????????????', case_detail)
            schema_id, new_data = updschema_data(data)
            new_url = url.format(schema_id)
            print(new_url)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data:", new_data)
            response = requests.put(url=new_url, headers=headers, data=new_data)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(table_sheet_name, row, column, response.status_code)
            write_result(table_sheet_name, row, column+4, response.text)
        elif case_detail == '???????????????':
            print('???????????????', case_detail)
            dss_id, new_data = upddss_data(data)
            new_url = url.format(dss_id)
            print(new_url)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data:", new_data)
            response = requests.put(url=new_url, headers=headers, data=new_data)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(table_sheet_name, row, column, response.status_code)
            write_result(table_sheet_name, row, column+4, response.text)
        elif case_detail == '??????dataset':
            print('???????????????', case_detail)
            dataset_id, new_data = upddataset_data(data)
            new_url = url.format(dataset_id)
            print(new_url)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data:", new_data)
            response = requests.put(url=new_url, headers=headers, data=new_data)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(table_sheet_name, row, column, response.status_code)
            write_result(table_sheet_name, row, column+4, response.text)
        elif '??????flow'in case_detail:
            print('???????????????', case_detail)
            dataset_id, new_data = update_flow_data(data)
            new_url = url.format(dataset_id)
            print(new_url)
            new_data = json.dumps(new_data, separators=(',', ':'))
            print("new_data:", new_data)
            response = requests.put(url=new_url, headers=headers, data=new_data)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(table_sheet_name, row, column, response.status_code)
            write_result(table_sheet_name, row, column+4, response.text)
        elif '????????????' in case_detail:
            types=case_detail.split("_")[1]
            print('???????????????', case_detail)
            url = url.format(data)
            tag_data_result=tag_data(types,data)
            response = requests.put(url=url, headers=headers, json=tag_data_result)
            print("response data:", response.status_code, response.text)
            clean_vaule(table_sheet_name, row, column)
            write_result(table_sheet_name, row, column, response.status_code)
            write_result(table_sheet_name, row, column+4, response.text)
        elif case_detail == '????????????':
            log.info("request   url???%s" % url)
            new_data = enable_role(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.put(url=url, headers=headers, data=new_data)
            log.info("response data???%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '????????????':
            log.info("request   url???%s" % url)
            new_data = enable_role(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.put(url=url, headers=headers, data=new_data)
            log.info("response data???%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '????????????':
            log.info("request   url???%s" % url)
            new_data = enable_user(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.put(url=url, headers=headers, data=new_data)
            log.info("response data???%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '????????????':
            log.info("request   url???%s" % url)
            new_data = enable_user(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.put(url=url, headers=headers, data=new_data)
            log.info("response data???%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '&' in str(data):
            # ????????????
            parameters = data.split('&')
            # ??????URL
            new_url = url.format(parameters[0])
            print(new_url)
            print(parameters)
            # ??????????????????
            parameters_data = parameters[1]
            if parameters_data.startswith('{'):
                response = requests.put(url=new_url, headers=headers, json=dict_res(parameters_data))
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(table_sheet_name, row, column, response.status_code)
                write_result(table_sheet_name, row, column+4, response.text)
            else:
                print('????????????%d???parameters?????????update?????????????????????id&{data}' % row)
        else:
            if data.startswith('select id'):
                result = ms.ExecuQuery(data)
                new_data = result[0]["id"]
                print(new_data, type(new_data))
                new_url = url.format(new_data)
                print('new_url:', new_url)
                response = requests.put(url=new_url, headers=headers)
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(table_sheet_name, row, column, response.status_code)
                write_result(table_sheet_name, row, column + 4, response.text)
            elif data.startswith('{') and data.endswith('}'):
                print(data)
                response = requests.put(url=url, headers=headers, data=data.encode('utf-8'))
                print("response data:", response.status_code, response.text)
                print(response.url, response.content)
                clean_vaule(table_sheet_name, row, column)
                write_result(table_sheet_name, row, column, response.status_code)
                write_result(table_sheet_name, row, column + 4, response.text)
            elif data.startswith('[') and data.endswith(']'):
                pass
            else:
                new_url = url.format(data)
                # print('new_url:', new_url)
                response = requests.put(url=new_url, headers=headers)
                print("response data:", response.status_code, response.text)
                clean_vaule(table_sheet_name, row, column)
                write_result(table_sheet_name, row, column, response.status_code)
                write_result(table_sheet_name, row, column + 4, response.text)
    except Exception as e:
        log.error("{}?????????????????????{}".format(case_detail,e))
        clean_vaule(table_sheet_name, row, column)
        write_result(table_sheet_name, row, column, '-1')
        write_result(table_sheet_name, row, column+4,  value='{"id":"-1"}')
    #else:
    #print('???%s?????????????????????????????????' % row)


def delete_request_result_check(url, host, data, table_sheet_name, row, column, headers):
    case_detail = case_table_sheet.cell(row=row, column=2).value
    log.info("???????????????%s" % case_detail)
    try:
        if isinstance(data, str):
            if case_detail == '':
                pass
            elif ("????????????") in case_detail:
                para=data.split("&")
                es_id=get_es_data(para[0],para[1],para[2],eval(para[3]))
                data={para[4]:es_id}
                response = requests.delete(url=url, headers=headers, json=data)
                print(response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
            else:
                if data.startswith('select id'):  # sql?????????????????????????????????
                    data_select_result = ms.ExecuQuery(data.encode('utf-8'))
                    #print(data_select_result)
                    #print(type(data_select_result))
                    datas = []
                    if data_select_result:
                        try:
                            for i in range(len(data_select_result)):
                                datas.append(data_select_result[i]["id"])
                        except:
                            print('????????????%d???SQL??????' % row)
                        else:
                            if len(datas) == 1:
                                # print(datas)
                                new_url = url.format(datas[0])
                                response = requests.delete(url=new_url, headers=headers)
                                print(response.url, response.status_code)
                                # ????????????status_code???response.text???????????????10?????????14???
                                clean_vaule(table_sheet_name, row, column)
                                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                            else:
                                print('????????? select ??????????????????????????????????????????')
                    else:
                        print('???%d????????????????????????' % row)
                    # ??????????????????????????????{"id":"7135cf6e-2b12-4282-90c4-bed9e2097d57","name":"gbj_for_jdbcDatasource_create_0301_1_0688","creator":"admin"}

                else:
                    new_url = url.format(data)
                    #print(new_url)
                    response = requests.delete(url=new_url, headers=headers)
                    print("response data:", response.status_code, response.text)
                    # ????????????status_code???response.text???????????????10?????????14???
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail=='??????????????????????????????':
            print(data)
            response = requests.delete(url=url, headers=headers,json=data)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif ("????????????????????????") in case_detail:
            if "_MINIO" in case_detail:
                data={"conf":{"password":"inforefiner","port":"9000","host":"192.168.1.81","region":"","username":"minio"},"name":minio_data}
            elif "_OZONE" in case_detail:
                data={"conf":{},"name":minio_data}
            response = requests.delete(url=url, headers=headers, json=data)
            print(response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif isinstance(data,list):
            response = requests.delete(url=url, headers=headers,data=json.dumps(data))
            print(response.url, response.status_code)
            # ????????????status_code???response.text???????????????10?????????14???
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        else:
            print('????????????%d??????data??????' % row)
    except Exception as e:
        log.error("{}?????????????????????{}".format(case_detail,e))
        clean_vaule(table_sheet_name, row, column)
        write_result(sheet=table_sheet_name, row=row, column=column, value='-1')
        write_result(sheet=table_sheet_name, row=row, column=column + 4, value='{"id":"-1"}')


#  ??????????????????
def write_result(sheet, row, column, value):
    sheet.cell(row=row, column=column, value=value)


#  ?????????????????????????????????????????????????????????
def clean_vaule(sheet, row, column):
    sheet.cell(row=row, column=column, value='')
    sheet.cell(row=row, column=column+1, value='')
    sheet.cell(row=row, column=column + 4, value='')
    sheet.cell(row=row, column=column + 5, value='')
    sheet.cell(row=row, column=column + 6, value='')
    sheet.cell(row=row, column=column + 7, value='')


# ??????code???text
class CheckResult(unittest.TestCase):

    def compare_code_result(self):
        """1.????????????code????????????????????????status code"""
        for row in range(2, all_rows+1):
            # ??????status code???????????????status code
            ex_status_code = case_table_sheet.cell(row=row, column=7).value
            ac_status_code = case_table_sheet.cell(row=row, column=8).value
            # ????????????status code????????????
            if ex_status_code and ac_status_code != '':
                # code????????????pass
                if ex_status_code == ac_status_code:
                    case_table_sheet.cell(row=row, column=9, value='pass')
                else:
                    case_table_sheet.cell(row=row, column=9, value='fail') # code?????????????????????????????????????????????
                    print('???????????????%s, ???????????????%s' % (ex_status_code, ac_status_code))
            else:
                print('??? %d ??? status_code??????' % row)
        case_table.save(ab_dir('api_cases.xlsx'))

    # ????????????response??????????????????response.text???????????????????????????????????????????????????
    def compare_text_result(self):
        for row in range(2, all_rows+1):
            response_text = case_table_sheet.cell(row=row, column=12).value  # ???????????????response.text
            response_text_dict = dict_res(response_text)
            expect_text = case_table_sheet.cell(row=row, column=10).value  # ????????????
            key_word = case_table_sheet.cell(row=row, column=3).value  # ???????????????
            code_result = case_table_sheet.cell(row=row, column=9).value  # status_code????????????
            relation = case_table_sheet.cell(row=row, column=11).value  # ??????text???response.text?????????
            #  1.status_code ????????????pass?????????????????????response.text??????????????????,
            #  2.status_code ????????????fail??????????????????????????????fail
            if code_result == 'pass':
                if key_word in ('create', 'query', 'update', 'delete'):
                    self.assert_deal(key_word, relation, expect_text, response_text, response_text_dict, row, 13)
                else:
                    print('????????????%d??????key_word' % row)
            elif code_result == 'fail':
                # case ?????????
                case_table_sheet.cell(row=row, column=14, value='fail')
                # case????????????
                case_table_sheet.cell(row=row, column=15, value='status_code???????????????%s' % code_result)
            else:
                print('???????????? %d ??? status_code????????????' % row)

        case_table.save(ab_dir('api_cases.xlsx'))

    #  ??????expect_text, response_text????????????????????????, ?????????????????????????????????????????????
    def assert_deal(self, key_word, relation, expect_text, response_text, response_text_dict, row, column):
        if key_word == 'create':
            if relation == '=':   # ?????????id?????????????????????????????????id?????????id?????????36
                if isinstance(response_text_dict, dict):
                    if response_text_dict.get("id"):
                        # ???????????????????????? id????????????????????????id????????????????????????id????????????
                        try:
                            self.assertEqual(expect_text, len(response_text_dict['id']), '???%d??????response_text????????????????????????' % row)
                        except:
                            print('??? %d ??? response_text?????????id?????????id???????????????' % row)
                            case_table_sheet.cell(row=row, column=column, value='fail')
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass')
                    else:
                        try:
                            self.assertEqual(expect_text, response_text, '???%d??????response_text????????????????????????' % row)
                        except:
                            print('??? %d ??? response_text?????????text?????????' % row)
                            case_table_sheet.cell(row=row, column=column, value='fail')
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass')
                else:  # ???????????????id???????????????????????????????????????id????????????
                    try:
                        self.assertEqual(expect_text, len(response_text), '???%d??????response_text????????????????????????' % row)
                    except:
                        print('??? %d ??? response_text?????????text?????????' % row)
                        case_table_sheet.cell(row=row, column=column, value='fail')
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass')

            elif relation == 'in':  # ????????????????????????????????????????????????id???????????????expect_text?????????response_text???
                try:
                    # self.assertIsNotNone(response_text_dict.get("id"), '??? %d ??? response_text????????????id' % row)
                    self.assertIn(expect_text, response_text, '??? %d ??? expect_text??????????????????????????????response_text???' % row)
                except:
                    print('??? %d ??? expect_text???????????????response_text?????? ??????????????????' % row)
                    case_table_sheet.cell(row=row, column=column, value='fail')
                else:
                    case_table_sheet.cell(row=row, column=column, value='pass')
            else:
                print('???????????? %d ??? ??????expect_text???response_text???relatrion' % row)
                case_table_sheet.cell(row=row, column=column, value='???????????????text?????????response.text???relatrion')
        elif key_word in ('query', 'update', 'delete'):
            if relation == '=':
                compare_result = re.findall('[a-z0-9]{8}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{12}', '%s' % (response_text))
                response_text_list = []
                response_text_list.append(response_text)
                #print("555555555555-6666666",response_text_list,compare_result)
                # ????????????id ??????????????????????????????
                if compare_result == response_text_list:
                    try:
                        self.assertEqual(expect_text, response_text, '???%s???expect_text???response_text?????????' % row)
                    except:
                        case_table_sheet.cell(row=row, column=column, value='fail')
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass')
                # ????????????
                elif expect_text == None and response_text == "":
                    case_table_sheet.cell(row=row, column=column, value='pass')

                else:
                    try:
                        if expect_text.find("createTime")>=0:
                            json_obj_exp = json.loads(expect_text)
                            results_exp = jsonpath.jsonpath(json_obj_exp,"$.[*].id")
                            print("results",results_exp)
                            json_obj_res = json.loads(response_text)
                            results_res = jsonpath.jsonpath(json_obj_res,"$.[*].id")
                            self.assertEqual(results_exp, results_res, '???%s???results_exp???results_res?????????' % row)
                        else:
                            self.assertEqual(expect_text, response_text, '???%s???expect_text???response_text?????????' % row)
                    except:
                        case_table_sheet.cell(row=row, column=column, value='fail')
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass')

            elif relation == 'in':
                try:
                    self.assertIn(expect_text, response_text, '??? %d ??? expect_text??????????????????????????????response_text???' % row)
                except:
                    print('??? %d ??? expect_text???response_text???????????? ??????????????????' % row)
                    case_table_sheet.cell(row=row, column=column, value='fail')
                else:
                    case_table_sheet.cell(row=row, column=column, value='pass')
            else:
                print('???????????? %d ??? ??????expect_text???response_text???relatrion' % row)
                case_table_sheet.cell(row=row, column=column, value='???????????????text?????????response.text???relatrion')
        else:
            print('???????????? %d ??? ???key_word' % row)
        case_table.save(ab_dir('api_cases.xlsx'))
    # ??????case???????????????
    def deal_result(self):
        # ??????????????????
        # deal_request_method()
        # ??????code
        self.compare_code_result()
        # ??????text
        self.compare_text_result()
        # ??????code result???text result??????case????????????
        for row in range(2, all_rows + 1):
            status_code_result = case_table_sheet.cell(row=row, column=9).value
            response_text_result = case_table_sheet.cell(row=row, column=13).value
            if status_code_result == 'pass' and response_text_result == 'pass':
                # print('????????????:%s ????????????' % case_table_sheet.cell(row=row, column=3).value)
                case_table_sheet.cell(row=row, column=14, value='pass')
                case_table_sheet.cell(row=row, column=15, value='')
            #elif status_code_result == 'fail' and response_text_result == 'pass':
            #    case_table_sheet.cell(row=row, column=14, value='fail')
            #    case_table_sheet.cell(row=row, column=15, value='%s--->???????????????status code????????????,?????????%s,?????????%s'
            #                                                    % (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=7).value, case_table_sheet.cell(row=row, column=8).value))
            elif status_code_result == 'pass' and response_text_result == 'fail':
                case_table_sheet.cell(row=row, column=14, value='fail')
                case_table_sheet.cell(row=row, column=15, value='%s--->???????????????????????????????????????' %
                                                                (case_table_sheet.cell(row=row, column=2).value))
            #elif status_code_result == 'fail' and response_text_result == 'fail':
            elif status_code_result == 'fail':
                case_table_sheet.cell(row=row, column=14, value='fail')
                case_table_sheet.cell(row=row, column=15, value='%s--->???????????????status code????????????????????????????????????????????????<api_cases.xlsx>????????????????????????'
                                                                % (case_table_sheet.cell(row=row, column=2).value))
            else:
                print('?????????status code???response.text????????????')
        case_table.save(ab_dir('api_cases.xlsx'))